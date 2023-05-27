from calendar import c
from itertools import count
from turtle import left
from automationDataProtector import settings
from collectEmail.utils.Outlook import Outlook
from collectEmail.models import Email, ScheduleOrLink, UltimateVerification
from collectEmail.utils.FilterEnum import FilterEnum
from django.utils import timezone
from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.styles import PatternFill, Font
from itertools import chain
import pandas as pd
import environ
import os
import re


class MainProcessCollect():

    GET_ENV = environ.Env()
    Normall = None
    OUTLOOK = Outlook()

    def __init__(self):
        if self.Normall is None:
            self.start_collect()

    def start_collect(self):
        self.OUTLOOK.login()
        self.OUTLOOK.readFolders()
        ids = self.OUTLOOK.readAllIdByDate(days=2)
        mails = self.OUTLOOK.getMailByIdsAndFrom(ids)
        self.wait_more_emails(mails)

    """ first request create file to write items for Link """

    def saveFirstFile(name):
        """ Verify file exist """
        if not os.path.exists(name):
            """ Save Report file """
            workbook = Workbook()

            sheet = workbook.active
            sheet["A1"] = "Hello"
            sheet["B1"] = "World!"

            workbook.save(filename=name)
            workbook.close()

    """ wait more emails if 2 hours not passed, it's over check if more emails exits or not and send email with link """

    def wait_more_emails(self, mails):
        time = UltimateVerification.objects.all().first()

        if time is None and len(mails) > 0:
            self.process_mails(mails)
        elif time is not None:
            if self.check_hours_passed(time) and len(mails) == 0:
                self.send_report_link()
            elif not self.check_hours_passed(time):
                self.process_mails(mails, time)
            else:
                self.process_mails(mails, time)
                self.send_report_link()

    """ verify hour passed is 2 hours or not"""

    def check_hours_passed(self, time):
        new = time.comprovate + timezone.timedelta(hours=1)
        now = timezone.now()
        return now >= new

    def process_mails(self, mails, time=None):
        self.mails_len = len(mails)

        for key, mail in mails.items():
            self.save_time(time)
            self.key_ = key
            try:
                if mail['schedule'] is not None and mail['link'] is None:
                    link = self.get_email_saved()
                    if link is not None:
                        self.apply_filters(mail['schedule'], link.body)
                    else:
                        self.save_email_body(mail, 'schedule')
                        continue

                elif mail['schedule'] is None and mail['link'] is not None:
                    schedule = self.get_email_saved()
                    if schedule is not None:
                        self.apply_filters(schedule.body, mail['link'])
                    else:
                        self.save_email_body(mail, 'link')
                        continue
                else:
                    self.apply_filters(mail['schedule'], mail['link'])
            except Exception as e:
                print(e)
                print('Error in process %s' % key)
                continue

        print('End process mails')

    """ send email with link and delete all sheets in excel file """

    def send_report_link(self):
        if self.Normall:
            return

        email = Email.objects.all().first()

        if email is not None:

            self.delete_sheet_default()

            state_send = self.OUTLOOK.send_mail(
                to=email.email, subject='Reportes para hacer el LINK',)
            if state_send:

                """ delete all files .xlsx """
                self.delete_files()

                """ delete time saved """
                time = UltimateVerification.objects.all().first()
                time.delete()

    print('function_send email with link')

    """ save time if time to verify is time passed 2 hours or send email with link """

    def save_time(self, time=None):
        if time is None:
            time = UltimateVerification.objects.all().first()
            if time is None:
                time = UltimateVerification.objects.create()
            else:
                time = timezone.now()
            time.save()
        else:
            if self.mails_len > 0:
                time.comprovate = timezone.now()
                time.save()

    """ Consult email body with id server  """

    def get_email_saved(self):
        return ScheduleOrLink.objects.filter(id=self.key_).first()

    """ Save email body when schedule or link is None for after process """

    def save_email_body(self, mail, type):
        ScheduleOrLink.objects.create(
            id=self.key_, body=mail[type].as_string(), type=type).save()

    """ Apply filters to schedule and link and write excel"""

    def apply_filters(self, schedule, link):
        schedule_ = schedule if isinstance(
            schedule, str) else schedule.as_string()
        link_ = link if isinstance(link, str) else link.as_string()

        self.table_schedule = pd.read_html(schedule_)
        self.table_link = pd.read_html(link_)

        """ remove characteres special from column name"""

        for column in self.table_schedule[0].columns:
            self.table_schedule[0].rename(
                columns={column: column.replace('= ', '')}, inplace=True)

        for column in self.table_link[0].columns:
            self.table_link[0].rename(
                columns={column: column.replace('= ', '')}, inplace=True)

        self.table_schedule[0].sort_values(by=['Specification'], inplace=True)
        self.table_link[0].sort_values(by=['Specification'], inplace=True)

        for f in FilterEnum.MALFORMED:
            self.replace_string_into_column(f, '', 'Specification')
            self.replace_string_into_column(f, '', 'Group')
            self.replace_string_into_column(f, '', 'Session Type')
            self.replace_string_into_column(f, '', 'Mode')
            self.replace_string_into_column(f, '', 'Start Time')
            self.replace_string_into_column(f, '', 'Files')
            self.replace_string_into_column(f, '', 'Session ID')
            self.replace_string_into_column(f, '', 'Next Execution')
            self.replace_string_into_column(f, '', 'Queuing')
            self.replace_string_into_column(f, '', 'Duration')
            self.replace_string_into_column(f, '', '# Media')
            self.replace_string_into_column(f, '', 'GB Written')
            self.replace_string_into_column(f, '', '# Errors')
            self.replace_string_into_column(f, '', '# Files')

        self.check_query_and_apply_filter("Specification.isna()")

        for f in FilterEnum.SPECIFICATION:
            self.check_query_and_apply_filter(
                "Specification.str.contains('(?i)"+f+"')")

        self.check_query_and_apply_filter("Group.isna()")

        for f in FilterEnum.GROUP:
            self.check_query_and_apply_filter(
                "Group.str.contains('(?i)"+f+"')")

        self.check_query_and_apply_filter("Group.str.len() < 1")
        self.check_query_and_apply_filter(
            "`Session Type`.str.contains('(?i)copy')")

        self.replace_string_start_column('(?i)mssql', '', 'Specification')
        self.replace_string_start_column('(?i)veagent', '', 'Specification')
        self.replace_string_start_column('(?i)hana', '', 'Specification', 1)
        self.replace_string_start_column('(?i)sap', '', 'Specification', 1)
        self.replace_string_start_column('(?i)oracle8', '', 'Specification')
        self.replace_string_start_column('(?i)idb', '', 'Specification', 1)
        self.replace_string_start_column('(?i)sybase', '', 'Specification', 1)

        for f in FilterEnum.STATUS:
            self.delete_schedule_items_with_status(
                "Status.str.contains('(?i)"+f+"')")

        self.delete_duplicates_items_from_schedules()
        self.remove_schedule_olds()
        self.copy_schedule_link_to_table_schedule()
        self.separate_table_schedule_men_sem()
        self.delete_email_saved()

        print('Filters applied %s' % self.key_)

    def process_veem_files(self, xlsx_file: Workbook, vcenter: str):
        df_1 = pd.DataFrame(columns=['Session Type', 'Specification',
                                     'Status', 'Mode', 'Start Time', 'Duration', 'GB Written'])

        df_2 = pd.DataFrame(columns=['Status', 'Objects Processed', 'Backup Type', 'Start Time', 'Duration',
                                     'Processing Rate (MB/Sec)', 'Data Size (GB)', 'Transferred (GB)', 'Total Backup Size (GB)'])

        """ select sheet to xlsx_file is 'Sheet2' """
        wb = xlsx_file['Sheet2']
        self.key_ = vcenter.split('_')[0]
        sheet_merged_cells_ = wb.merged_cells.ranges
        sheet_merged_cells_sorted = sorted(
            sheet_merged_cells_, key=lambda x: x.min_row)

        """ Filter merged cell by sem or men """
        sheet_merged_cells_sem = list(filter(lambda x: x.min_row > 7 and 'sem' in wb.cell(
            row=x.min_row, column=x.min_col).value.lower(), sheet_merged_cells_sorted))
        sheet_merged_cells_men = list(filter(lambda x: x.min_row > 7 and 'men' in wb.cell(
            row=x.min_row, column=x.min_col).value.lower(), sheet_merged_cells_sorted))
        sheet_merged_cells_delete = list(filter(
            lambda x: x.min_row > 7 and x not in sheet_merged_cells_sem and x not in sheet_merged_cells_men, sheet_merged_cells_sorted))

        session_type = wb.cell(
            sheet_merged_cells_sorted[1].min_row, sheet_merged_cells_sorted[1].min_col).value.split(':')[1].strip()

        for index, cell in enumerate(sheet_merged_cells_sorted):

            if index+1 < len(sheet_merged_cells_sorted):
                cell_end = sheet_merged_cells_sorted[index+1]
                [range_start, range_end] = [cell.min_row+1, cell_end.min_row-1]
            else:
                [range_start, range_end] = [cell.min_row+1, wb.max_row]

            if cell in chain(sheet_merged_cells_sem, sheet_merged_cells_men, sheet_merged_cells_delete):
                specification = wb.cell(
                    row=cell.min_row, column=cell.min_col).value.split(':')[1].strip()

            if cell in sheet_merged_cells_delete:
                df_2 = pd.concat([df_2, pd.DataFrame(
                    {'Status': [specification]})], axis=0, ignore_index=True)

                for row in range(range_start, range_end+1):
                    status = wb.cell(row=row, column=cell.min_col).value
                    process = wb.cell(row=row, column=cell.min_col+1).value
                    type = wb.cell(row=row, column=cell.min_col+2).value
                    start_time = wb.cell(row=row, column=cell.min_col+3).value
                    duration = wb.cell(row=row, column=cell.min_col+4).value
                    rate = wb.cell(row=row, column=cell.min_col+5).value
                    data_size = wb.cell(row=row, column=cell.min_col+6).value
                    transferred = wb.cell(row=row, column=cell.min_col+7).value
                    total_backup = wb.cell(
                        row=row, column=cell.min_col+8).value

                    df_2 = pd.concat([df_2, pd.DataFrame({'Status': [status], 'Objects Processed': [process], 'Backup Type': [type], 'Start Time': [start_time], 'Duration': [duration],
                                                          'Processing Rate (MB/Sec)': [rate], 'Data Size (GB)': [data_size], 'Transferred (GB)': [transferred], 'Total Backup Size (GB)': [total_backup]})], axis=0, ignore_index=True)
            elif cell in chain(sheet_merged_cells_sem, sheet_merged_cells_men):
                df = pd.DataFrame(columns=['Session Type', 'Specification',
                                           'Status', 'Mode', 'Start Time', 'Duration', 'GB Written'])

                for row in range(range_start, range_end+1):
                    status = wb.cell(row=row, column=cell.min_col).value
                    mode = wb.cell(row=row, column=cell.min_col+2).value
                    start_time = wb.cell(row=row, column=cell.min_col+3).value
                    duration = wb.cell(row=row, column=cell.min_col+4).value
                    gb_written = wb.cell(row=row, column=cell.min_col+8).value

                    df = pd.concat([df, pd.DataFrame({'Session Type': [session_type], 'Specification': [specification], 'Status': [status], 'Mode': [mode],
                                                      'Start Time': [start_time], 'Duration': [duration], 'GB Written': [gb_written]})], axis=0, ignore_index=True)
                """ Sort by Start Time and only get two ultimates when Mode equal [Incremental,Synthetic Full]"""
                df = df.sort_values(by=['Start Time'], ascending=True)
                df = df[df['Mode'].isin(['Incremental', 'Synthetic Full'])]
                """ detect if Mode only contain Incremental | Synthetic Full """
                if len(df['Mode'].unique()) == 2:
                    """ if mode unique > 2 get the last 2 equals to [Incremental,Synthetic Full]"""
                    df = df.groupby(['Mode'], as_index=False).max()
                    df = df.reset_index(drop=True)

                else:
                    df = df.iloc[-1:]
                df = df.reset_index(drop=True)
                df_1 = pd.concat([df_1, df])

        self.write_excel_file(df_1, self.GET_ENV('FILE_7'))
        self.write_excel_file(df_2, self.GET_ENV('FILE_6'))

    def process_pbi_files(self, xlsx_file: Workbook, file_name: str):
        """ convert woorbook to dataframe + headers """
        df = pd.read_excel(xlsx_file, header=0, engine='openpyxl')

        """ delete colums not needed """
        df.drop(FilterEnum.PBI, axis=1, inplace=True)

        """ delete job_name if bit content 'COPIA,copia,Copia' """
        df.query("job_name.str.contains('COPIA|copia|Copia') == False", inplace=True)

        """ order by start_time """
        df.sort_values(by=['start_time'], ascending=False, inplace=True)

        """ Delete duplicated values """
        df.drop_duplicates(subset=['job_name'], keep='first', inplace=True)

        """ Order by job_name """
        df.sort_values(by=['job_name'], ascending=True, inplace=True)

        """ Rename file PBI_.xlsx to  file_name"""
        os.rename(self.GET_ENV('FILE_8'), file_name)

        self.key_ = xlsx_file.active.title
        self.write_excel_file(df, file_name)

        woorkbook = load_workbook(file_name)
        sheet = woorkbook['Export']
        headers_range = sheet['A1':'J1']
        fill = PatternFill(start_color='ffd100',
                           end_color='ffd100', fill_type='solid')

        for cell in headers_range[0]:
            cell.fill = fill

        font = Font(bold=True)
        for cell in headers_range[0]:
            cell.font = font

        woorkbook.save(file_name)

    def delete_email_saved(self):
        ScheduleOrLink.objects.filter(id=self.key_).delete()

    """ check if specification contains dia if true delete row """

    def check_query_and_apply_filter(self, query):
        [schedule_colum_exist, link_colum_exist] = self.colum_exist(query)

        """ Saved deleted rows for vcenter """
        table_schedule_deleted_days = self.table_schedule[0].query(
            query, engine='python') if schedule_colum_exist else None

        table_link_deleted_days = self.table_link[0].query(
            query, engine='python') if link_colum_exist else None

        """ write excel with deleted rows """
        if table_schedule_deleted_days is not None:
            self.write_excel_file(
                table_schedule_deleted_days, self.GET_ENV('FILE_2'))

        if table_link_deleted_days is not None:
            self.write_excel_file(table_link_deleted_days,
                                  self.GET_ENV('FILE_3'))

        """ Delete rows with dia in schedule or link """
        if schedule_colum_exist:
            self.table_schedule[0].drop(
                table_schedule_deleted_days.index, inplace=True)
            self.table_schedule[0].reset_index(drop=True, inplace=True)
        if link_colum_exist:
            self.table_link[0].drop(
                table_link_deleted_days.index, inplace=True)
            self.table_link[0].reset_index(drop=True, inplace=True)

    """ replace into colum values start with, for new values """

    def replace_string_start_column(self, old, new, colum, limit=0):
        start_with = old.replace('(?i)', '')
        self.table_schedule[0][colum] = self.table_schedule[0][colum].map(lambda x: re.sub(
            old, new, x, limit) if x.lower().startswith(start_with) else x)

        self.table_schedule[0][colum] = self.table_schedule[0][colum].str.strip()

    """ replace into colum values for new values """

    def replace_string_into_column(self, old, new, colum):
        [schedule_colum_exist, link_colum_exist] = self.colum_exist(colum)

        if schedule_colum_exist:
            if pd.api.types.infer_dtype(self.table_schedule[0][colum]) == 'string':
                self.table_schedule[0][colum] = self.table_schedule[0][colum].str.replace(
                    old, new)
        if link_colum_exist:
            if pd.api.types.infer_dtype(self.table_link[0][colum]) == 'string':
                self.table_link[0][colum] = self.table_link[0][colum].str.replace(
                    old, new)

    """ Return if colum exist in table """

    def colum_exist(self, query):
        first_value_split = query.split('.')[0].replace("`", "")
        schedule_colum_exist = first_value_split in self.table_schedule[0].columns.values
        link_colum_exist = first_value_split in self.table_link[0].columns.values
        return [schedule_colum_exist, link_colum_exist]

    """ Delete items with status """

    def delete_schedule_items_with_status(self, query):
        table_only_filters = self.table_schedule[0].query(
            query, engine='python')

        self.write_excel_file(table_only_filters, self.GET_ENV('FILE_2'))

        self.table_schedule[0].drop(table_only_filters.index, inplace=True)

    """ Delete duplicate items in colum """

    def delete_duplicates_items_from_schedules(self):
        self.table_schedule[0].sort_values(
            by=['Start Time'], inplace=True)

        self.table_schedule[0].drop_duplicates(
            subset=['Specification'], keep='last', inplace=True)

        self.table_schedule[0].sort_values(by=['Specification'], inplace=True)
        self.table_schedule[0].reset_index(drop=True, inplace=True)
        self.table_link[0].reset_index(drop=True, inplace=True)

    """ Remove old schedule not contain in link """

    def remove_schedule_olds(self):
        self.table_schedule[0]['Specification_lower'] = self.table_schedule[0]['Specification'].str.lower()
        self.table_link[0]['Specification_lower'] = self.table_link[0]['Specification'].str.lower()

        table_schedule_olds = pd.merge(self.table_link[0], self.table_schedule[0],
                                       how='right', on=['Specification_lower'], indicator=True)

        table_schedule_olds = table_schedule_olds.query(
            '_merge == "left_only"', engine='python')

        self.write_excel_file(table_schedule_olds, self.GET_ENV('FILE_2'))

        table_schedule_loc = self.table_schedule[0].loc[self.table_schedule[0]['Specification_lower'].isin(
            table_schedule_olds['Specification_lower'])]

        """ delete only table_schedule._merge == True """
        self.table_schedule[0].drop(table_schedule_loc.index, inplace=True)

    """ Copy schedule link to table schedule """

    def copy_schedule_link_to_table_schedule(self):
        table_schedule_copy = pd.merge(self.table_schedule[0], self.table_link[0], how='right', on=[
            'Specification_lower'], indicator=True).set_axis(self.table_link[0].index)

        table_schedule_copy_both = table_schedule_copy.query(
            '_merge == "both"', engine='python')

        self.update_table_schedule(table_schedule_copy_both)

        """ delete column innecesary """
        self.table_link[0].drop(
            ['Type', 'Specification_lower'], axis=1, inplace=True)
        self.table_schedule[0].drop(
            ['Specification_lower', 'Success', '# Warnings'], axis=1, inplace=True)

        table_schedule_copy_right = table_schedule_copy.query(
            '_merge == "right_only"', engine='python')

        self.table_schedule[0] = pd.concat(
            [self.table_schedule[0], self.table_link[0].loc[table_schedule_copy_right.index]], ignore_index=True)
    """ Update Colums in table_schedule """

    def update_table_schedule(self, dataframe):
        """ Add new columns """
        self.table_schedule[0]['Group'] = '-'
        self.table_schedule[0]['Next Execution'] = '-'

        """ Update data from column if Specification_lower equals in both tables """

        for _, row in dataframe.iterrows():
            self.table_schedule[0].loc[self.table_schedule[0]['Specification_lower'] == row['Specification_lower'], [
                'Group', 'Next Execution']] = [row['Group'], row['Next Execution']]

    """ Separate schedule to week and month """

    def separate_table_schedule_men_sem(self):

        table_schedule_sem = self.table_schedule[0].query(
            "Specification.str.contains('(?i)sem')", engine='python')
        table_schedule_men = self.table_schedule[0].query(
            "Specification.str.contains('(?i)men')")

        self.table_schedule[0].drop(table_schedule_sem.index, inplace=True)
        self.table_schedule[0].drop(table_schedule_men.index, inplace=True)

        table_schedule_sem = table_schedule_sem.sort_values(
            by=['Specification'])
        table_schedule_men = table_schedule_men.sort_values(
            by=['Specification'])
        self.table_schedule[0].sort_values(by=['Specification'], inplace=True)

        self.write_excel_file(table_schedule_sem, self.GET_ENV('FILE_1'))
        self.write_excel_file(table_schedule_men, self.GET_ENV('FILE_4'))
        self.write_excel_file(self.table_schedule[0], self.GET_ENV('FILE_5'))

    """ Write excel file """

    def write_excel_file(self, table, file_name):
        wb = load_workbook(file_name)
        sheet_exist = self.key_ in wb.sheetnames

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            writer.book = wb

            """ Copy all sheets of file """
            writer.sheets.update(dict((ws.title, ws)
                                      for ws in wb.worksheets))

            if sheet_exist:
                """ writer in sheet if exist"""
                book_writed = len(wb[self.key_]['A']) + 1

                table.to_excel(writer, sheet_name=self.key_,
                               startrow=book_writed, index=False, header=False)
            else:
                table.to_excel(writer, sheet_name=self.key_,
                               index=False)

    """ Delete sheet innecesary"""

    def deleteSheet(self, file_name, sheet_name):
        wb_f = load_workbook(file_name)
        if len(wb_f.sheetnames) > 1 and sheet_name in wb_f.sheetnames:
            del wb_f[sheet_name]
            wb_f.save(file_name)
            wb_f.close()

    def delete_files(self):
        file = [self.GET_ENV('FILE_1'), self.GET_ENV('FILE_2'), self.GET_ENV(
            'FILE_3'), self.GET_ENV('FILE_4'), self.GET_ENV('FILE_5'), self.GET_ENV('FILE_6'), self.GET_ENV('FILE_7')]
        for f in file:
            os.remove(os.path.join(settings.BASE_DIR, f))

        for f in os.listdir(settings.BASE_DIR):
            if f.startswith('PBI_'):
                os.remove(os.path.join(settings.BASE_DIR, f))

        file.append(self.GET_ENV('FILE_8'))

        for f in file:
            MainProcessCollect.saveFirstFile(f)

    def delete_sheet_default(self):
        for i in range(1, 8):
            self.deleteSheet(self.GET_ENV('FILE_%s' % i), 'Sheet')

        """ verify if exist file start with PBI_ """
        for file in os.listdir(settings.BASE_DIR):
            if file.startswith('PBI_'):
                self.deleteSheet(file, 'Sheet')
