from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from os.path import basename
from urllib.parse import urlencode
from openpyxl import load_workbook
from automationDataProtector import settings
import os
import datetime
import email
import imaplib
import smtplib
import environ
import requests


class Outlook():

    LOG_INFO = "OUTLOOK: "
    GET_ENV = environ.Env()

    def access_token(self):
        tenant = self.GET_ENV('TENANT_ID')
        client_id = self.GET_ENV('CLIENT_ID')
        client_secret = self.GET_ENV('CLIENT_SECRET')

        url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
        payload = urlencode({
            "client_id": client_id,
            "scope": "https://outlook.office365.com/.default",
            "client_secret": client_secret,
            "grant_type": "client_credentials"
        })
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        return requests.post(url, headers=headers, data=payload).json()['access_token']

    def generate_auth_token(self, token):
        return f"user={self.email}\1auth=Bearer {token}\1\1".encode()

    def login(self):
        self.email = self.GET_ENV('EMAIL')
        self.password = self.GET_ENV('PASSWORD')
        loggin_attempt = 0

        while True:
            try:
                self.imap = imaplib.IMAP4_SSL(self.GET_ENV(
                    'IMAP_SERVER'), self.GET_ENV('IMAP_PORT'))

                token = self.access_token()

                r, d = self.imap.authenticate(
                    'XOAUTH2', lambda _: self.generate_auth_token(token))

                """ r, d = self.imap.login(self.email, self.password) """

                assert r == 'OK', 'login failed: %s' % str(r)

                print(self.LOG_INFO + "Accediendo como %s" % email, d)
                return
            except Exception as e:
                print(self.LOG_INFO + "Error al iniciar sesion: %s" % str(e))

                loggin_attempt += 1
                if loggin_attempt > 3:
                    continue

                assert False, 'login failed'

    def send_mail(self, to, subject):

        msg = MIMEMultipart()
        msg['From'] = self.email
        msg['To'] = to
        msg['Subject'] = subject

        msg.attach(MIMEText('Reportes para generar el LINK'))

        """ add files to email """
        for i in range(1, 8):
            wb = load_workbook(self.GET_ENV('FILE_%s' % i))
            if 'Sheet' in wb.sheetnames:
                continue
            with open(self.GET_ENV('FILE_%s' % i), 'rb') as f:
                part = MIMEApplication(
                    f.read(), Name=basename(self.GET_ENV('FILE_%s' % i)))
            part['Content-Disposition'] = 'attachment; filename="%s"' % basename(
                self.GET_ENV('FILE_%s' % i))
            msg.attach(part)

        for file in os.listdir(settings.BASE_DIR):
            if file.startswith('PBI_'):
                wb = load_workbook(file)
                if 'Sheet' in wb.sheetnames:
                    continue
                with open(file, 'rb') as f:
                    part = MIMEApplication(f.read(), Name=basename(file))
                part['Content-Disposition'] = 'attachment; filename="%s"' % basename(
                    file)
                msg.attach(part)

        try:
            self.smtp = smtplib.SMTP(self.GET_ENV(
                'SMTP_SERVER'), self.GET_ENV('SMTP_PORT'))
            self.smtp.ehlo()
            self.smtp.starttls()
            self.smtp.login(self.email, self.password)
            self.smtp.sendmail(msg['From'], msg['To'], msg.as_string())
        except Exception as e:
            print(self.LOG_INFO + "Error al enviar el correo: %s" % str(e))
            return False

        return True

    def readFolders(self, folder='Inbox'):
        return self.imap.select(folder)

    def sinceDate(self, days):
        date = datetime.datetime.now() - datetime.timedelta(days=days)
        return date.strftime('%d-%b-%Y')

    def readAllIdByDate(self, days=1):
        r, d = self.imap.search(
            None, '(SINCE "'+self.sinceDate(days)+'")', 'UNSEEN')
        return d[0].split(b' ')

    def getEmail(self, id):
        r, d = self.imap.fetch(id, '(RFC822)')
        assert r == 'OK', 'fetch failed: %s' % str(r)
        self.email_message = email.message_from_bytes(d[0][1])
        return

    def getBody(self):
        if self.email_message.is_multipart():

            html = None
            for part in self.email_message.walk():
                if part.get_filename() is None:
                    continue
                html = part.get_payload(decode=True).decode(
                    encoding='utf-8', errors='ignore')
            body = {
                'message': self.email_message.get_payload(0) if html is None else html
            }
            return body
        else:
            body = {
                'message': self.email_message.get_payload(0)
            }
            return body

    def getMailByIdsAndFrom(self, ids):

        stack = {}

        if ids[0] == b'':
            return stack

        for id in ids:
            self.getEmail(id)

            if self.email_message['To'] is None:
                continue

            if self.GET_ENV('EMAIL') in self.email_message['To'] and ('link' in self.email_message['Subject'].lower() or 'schedule' in self.email_message['Subject'].lower()):
                subject = self.email_message['Subject'].replace('-', '')
                split_subject = subject.split(' ')

                if "" in split_subject:
                    split_subject.remove("")

                vcenter = split_subject[len(split_subject)-1]
                _type = 'schedule' if 'schedule' in [
                    x.lower() for x in split_subject] else 'link'

                if vcenter in stack:
                    stack[vcenter] = {
                        'schedule': self.getBody()['message'] if _type == 'schedule' else stack[vcenter]['schedule'],
                        'link': self.getBody()['message'] if _type == 'link' else stack[vcenter]['link']
                    }
                else:
                    stack[vcenter] = {
                        'schedule': self.getBody()['message'] if _type == 'schedule' else None,
                        'link': self.getBody()['message'] if _type == 'link' else None
                    }

        print(self.LOG_INFO + "Emails encontrados y por procesar: %s" % len(stack))
        return stack
