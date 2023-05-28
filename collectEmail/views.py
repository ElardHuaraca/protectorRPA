from django.shortcuts import render
from django.http import HttpResponseNotFound, JsonResponse, HttpResponse
from collectEmail.models import Email, UltimateVerification
from collectEmail.utils.Threads import ThreadsStart, MainProcessCollect
from django.utils import timezone
from openpyxl import load_workbook
from automationDataProtector import settings
import re
import os
import zipfile
import io
import environ

# Create your views here.

start = False


def index(request):
    global start

    context = {
        'error': False,
        'success': False,
        'message': '',
    }

    if not start:
        ThreadsStart()
        start = True

    return render(request, 'index.html', context)


def save(request):
    if (request.POST):
        email = request.POST.get('email')
        context = {}
        if (email == ''):
            context = {
                'error': True,
                'success': False,
                'message': 'El correo es requerido'
            }
        elif (isValidEmail(email)):
            Email.objects.all().delete()

            save_mail = Email.objects.create()
            save_mail.email = email
            save_mail.save()

            context = {
                'error': False,
                'success': True,
                'message': 'El correo se guardo correctamente'
            }
        else:
            context = {
                'error': True,
                'success': False,
                'message': 'El correo no es valido'
            }

        return render(request, 'index.html', context)
    else:
        return HttpResponseNotFound()


def download(_):
    GET_ENV = environ.Env()
    tempt_file = io.BytesIO()

    MainProcessCollect.Normall = True
    mainProcessCollect = MainProcessCollect()
    mainProcessCollect.delete_sheet_default()

    with zipfile.ZipFile(tempt_file, 'w') as zip:
        for file in os.listdir(settings.BASE_DIR):
            for i in range(1, 9):
                if file.startswith(GET_ENV('FILE_%s' % i).split('.')[0]):
                    zip.write(file)
    tempt_file.seek(0)

    response = HttpResponse(tempt_file, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=%s' % 'reportes.zip'

    deleteTimeOnDownloadReports()
    mainProcessCollect.delete_files()
    MainProcessCollect.Normall = None

    return response


def isValidEmail(email):
    """ verify is email is valid and contain domain canvia.com in expresion regular"""

    if (re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email)):
        if (re.search(r'canvia.com', email)):
            return True
    return False


def processFiles(request):
    files = os.listdir('temp')

    for file in files:
        os.remove('temp/%s' % file)

    for file in os.listdir(settings.BASE_DIR):
        if file.startswith('PBI_'):
            if re.match(r'^PBI_\d{8}\.xlsx', file) != None:
                os.remove(file)
                MainProcessCollect.saveFirstFile('PBI_.xlsx')

    list_files = request.FILES.getlist('files')

    if not os.path.exists('temp'):
        os.makedirs('temp')

    for file in list_files:
        with open('temp/%s' % file.name, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

    stack = {}
    files = os.listdir('temp')

    MainProcessCollect.Normall = True
    mainProcessCollect = MainProcessCollect()

    for _file in files:
        [vcenter, type, fileExtension] = getTypeFile(_file)

        if fileExtension != 'html':
            if _file.startswith('PBI_'):
                mainProcessCollect.process_pbi_files(
                    load_workbook('temp/%s' % _file), _file)
            else:
                mainProcessCollect.process_veem_files(
                    load_workbook('temp/%s' % _file), vcenter)
        elif _file.lower().startswith('schedule') or _file.lower().startswith('link'):
            """ open file and read all lines """
            file = open('temp/%s' % _file, 'r')
            html = file.read().splitlines()
            file.close()

            """ concat all string split in array to one string """
            html = ''.join(html)

            if vcenter in stack:
                stack[vcenter] = {
                    'schedule': html if type == 'schedule' else stack[vcenter]['schedule'],
                    'link': html if type == 'link' else stack[vcenter]['link']
                }
            else:
                stack[vcenter] = {
                    'schedule': html if type == 'schedule' else None,
                    'link': html if type == 'link' else None
                }
        else:
            continue

    mainProcessCollect.wait_more_emails(stack)
    MainProcessCollect.Normall = None

    return JsonResponse({'status': 'complete'})


def getTypeFile(filename):
    _name = str(filename).split('.')[0]
    fileExtension = str(filename).split('.')[-1]
    vcenter = _name.split(' ')[-1]
    type = 'schedule' if 'schedule' in _name.lower() else 'link'
    return [vcenter, type, fileExtension]


def updateTimeForceSendMail():
    time = UltimateVerification.objects.all().first()
    if not time:
        time = UltimateVerification.objects.create()
    time.comprovate = timezone.now() - timezone.timedelta(hours=4)
    time.save()


def deleteTimeOnDownloadReports():
    time = UltimateVerification.objects.all().first()
    if time:
        time.delete()
